import os
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO
from pathlib import Path

def export_to_excel(metrics, output_xlsx):
    status = {"WB": "OFF", "CLAHE": "OFF", "GAMMA": "OFF", "SHARPEN": "OFF"}
    status_file = Path("results/metrics/current_ablation_status.txt")
    if status_file.exists():
        with open(status_file, "r") as f:
            for line in f:
                k, v = line.strip().split(":")
                status[k] = "ON ✅" if v == "True" else "OFF ❌"

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Ablation Test Result"; ws.views.sheetView[0].showGridLines = True
    
    PRI_COLOR, ZEBRA_COLOR, BORDER_COLOR = "2C3E50", "F8F9F9", "BDC3C7"
    font_title = Font(name="Arial", size=14, bold=True, color=PRI_COLOR)
    font_section = Font(name="Arial", size=11, bold=True, color=PRI_COLOR)
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_body, font_body_bold = Font(name="Arial", size=10), Font(name="Arial", size=10, bold=True)
    
    fill_header = PatternFill(start_color=PRI_COLOR, end_color=PRI_COLOR, fill_type="solid")
    fill_zebra = PatternFill(start_color=ZEBRA_COLOR, end_color=ZEBRA_COLOR, fill_type="solid")
    border_all = Border(left=Side(style="thin", color=BORDER_COLOR), right=Side(style="thin", color=BORDER_COLOR), top=Side(style="thin", color=BORDER_COLOR), bottom=Side(style="thin", color=BORDER_COLOR))
    
    ws["B2"] = "Ablation Study Scientific Evaluation Report"
    ws["B2"].font = font_title
    
    ws["B4"] = "1. Image Enhancement Configuration"
    ws["B4"].font = font_section
    for idx, h in enumerate(["Technique Component", "Ablation Status"], start=2):
        c = ws.cell(row=5, column=idx, value=h); c.font = font_header; c.fill = fill_header; c.alignment = Alignment(horizontal="center"); c.border = border_all
        
    for idx, (k, v) in enumerate(status.items(), start=6):
        c1 = ws.cell(row=idx, column=2, value=k); c1.font = font_body; c1.border = border_all
        c2 = ws.cell(row=idx, column=3, value=v); c2.font = font_body_bold; c2.border = border_all; c2.alignment = Alignment(horizontal="center")
        if idx % 2 == 1: c1.fill = fill_zebra; c2.fill = fill_zebra

    ws["B12"] = "2. Final Performance Metrics (Evaluated strictly on TEST SET)"
    ws["B12"].font = font_section
    for idx, h in enumerate(["Metric", "Score Value (0-1)"], start=2):
        c = ws.cell(row=13, column=idx, value=h); c.font = font_header; c.fill = fill_header; c.alignment = Alignment(horizontal="center"); c.border = border_all
        
    m_data = [("mAP50", metrics["map50"]), ("mAP50-95", metrics["map95"]), ("Precision", metrics["precision"]), ("Recall", metrics["recall"])]
    for idx, (k, v) in enumerate(m_data, start=14):
        c1 = ws.cell(row=idx, column=2, value=k); c1.font = font_body; c1.border = border_all
        c2 = ws.cell(row=idx, column=3, value=v); c2.font = font_body_bold; c2.border = border_all; c2.alignment = Alignment(horizontal="right"); c2.number_format = "0.0000"
        if idx % 2 == 1: c1.fill = fill_zebra; c2.fill = fill_zebra

    for col in ws.columns:
        max_l = max(len(str(cell.value or '')) for cell in col if cell.row != 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_l + 6, 16)
        
    wb.save(output_xlsx)
    print(f"🥇 [DONE BƯỚC 4] Báo cáo khoa học đã xuất bản tại: {output_xlsx}")

def main():
    parser = argparse.ArgumentParser(description="Step 4: Evaluation on TEST set only")
    parser.add_argument("--weights", type=str, required=True, help="Đường dẫn file .pt muốn đem ra chấm điểm")
    parser.add_argument("--data", type=str, default="configs/scoral.yaml")
    args = parser.parse_args()

    print("\n=================================================================")
    print(f"🎯 TIẾN HÀNH KIỂM THỬ SÒNG PHẲNG TRÊN TẬP TEST XA LẠ")
    print(f"   Model sử dụng: {Path(args.weights).name}")
    print("=================================================================")
    
    model = YOLO(args.weights)
    metrics_obj = model.val(data=args.data, imgsz=640, split='test', project="results", name="test_eval", verbose=False)
    
    metrics_dict = {
        "map50": float(metrics_obj.box.map50),
        "map95": float(metrics_obj.box.map),
        "precision": float(metrics_obj.box.mp),
        "recall": float(metrics_obj.box.mr)
    }

    export_to_excel(metrics_dict, "results/metrics/ablation_summary.xlsx")

if __name__ == "__main__":
    main()